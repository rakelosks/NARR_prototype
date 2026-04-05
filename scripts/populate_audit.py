#!/usr/bin/env python3
"""
Pre-populate corpus/rq1_accuracy_audit.xlsx from corpus run JSON files.

Reads run_*_query.json, extracts narrative sentences as claims, and matches
them only to structured fields from response["evidence_bundle"] (metrics,
rankings, trends, etc.). key_findings inside the bundle are excluded from
matching to avoid comparing narrative claims to the same narrative digest.
Does not use Vega rows or a narrative self-digest. When a token/string match
hits the bundle, fills Verdict (Accurate) and a short Note for manual review.

Marks preview (non-LLM) runs on Pipeline Success and omits their timing from
Narrative Summary column N. Writes corpus/rq1_accuracy_audit_populated.xlsx.

Uses openpyxl so template formatting and data validations are preserved;
Claim Audit validation ranges are expanded so many claim rows stay valid.
"""

from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

REPO_ROOT = Path(__file__).resolve().parent.parent
CORPUS_DIR = REPO_ROOT / "corpus"
TEMPLATE_XLSX = CORPUS_DIR / "rq1_accuracy_audit.xlsx"
OUTPUT_XLSX = CORPUS_DIR / "rq1_accuracy_audit_populated.xlsx"

CLAIM_SHEET = "Claim Audit"
SUMMARY_SHEET = "Narrative Summary"
PIPELINE_SHEET = "Pipeline Success"

VERDICT_AUTO_MATCH = "Accurate"
NOTE_AUTO_MATCH = (
    "Auto: string/token overlap vs evidence_bundle metrics only — confirm accuracy."
)
NOTE_NO_BUNDLE = "No evidence_bundle in JSON — cannot match to structured metrics."

# --- Narrative extraction -------------------------------------------------

DIR_PATTERN = re.compile(
    r"\b(increased|decreases?|decreased|increasing|decreasing|increase|decrease|"
    r"higher|lower|highest|lowest|more|less|fewer|greater|smallest|largest|"
    r"grew|grown|growth|declined|declining|decline|rose|risen|rising|fallen|falling|"
    r"dropped|dropping|surged|shrunk|shrank|shrinking|expanding|worsening|improving|"
    r"stable|stagnant|flat|peaked|peak|trough|"
    r"upward|downward|trending)\b",
    re.I,
)

YEAR_PATTERN = re.compile(r"\b(19|20)\d{2}\b")
PCT_PATTERN = re.compile(r"\d+(?:\.\d+)?%")
NUMBER_PATTERN = re.compile(
    r"-?\d{1,3}(?:,\d{3})+(?:\.\d+)?|-?\d+(?:\.\d+)?",
)

HEDGE_PATTERN = re.compile(
    r"\b(?:approximately|approx\.?|roughly|about|around|~|ca\.|circa|"
    r"may|might|could|can|suggest(?:s|ed|ing)?|appear(?:s|ed)?|seem(?:s|ed)?|"
    r"likely|potentially?|probable|probably|estimated|uncertain|unclear|"
    r"perhaps|possibly|appears to be|tends to)\b",
    re.I,
)


def split_sentences(text: str) -> list[str]:
    if not text or not str(text).strip():
        return []
    parts = re.split(r"(?<=[.!?])\s+", str(text).strip())
    out = [p.strip() for p in parts if p.strip()]
    if len(out) == 1 and text.strip() and text.strip()[-1] not in ".!?":
        return [text.strip()]
    return out


def has_numerical_signal(s: str) -> bool:
    if YEAR_PATTERN.search(s) or PCT_PATTERN.search(s):
        return True
    if NUMBER_PATTERN.search(s):
        return True
    return False


def classify_claim_type(
    sentence: str,
    entities_lower: set[str],
    *,
    evidence_matched: bool,
) -> str:
    s_low = sentence.lower()
    if HEDGE_PATTERN.search(sentence):
        return "Hedging"
    has_num = has_numerical_signal(sentence)
    has_dir = bool(DIR_PATTERN.search(sentence))
    if not evidence_matched:
        if has_num or has_dir:
            return "Fabrication"
        for ent in entities_lower:
            if len(ent) >= 3 and ent in s_low:
                return "Fabrication"
        return ""
    if has_num:
        return "Numerical"
    if has_dir:
        return "Directional"
    for ent in entities_lower:
        if len(ent) >= 3 and ent in s_low:
            return "Entity"
    return ""


def extract_number_tokens(sentence: str) -> set[str]:
    """Normalized substrings to match inside metric blobs."""
    tokens: set[str] = set()
    for m in YEAR_PATTERN.finditer(sentence):
        tokens.add(m.group(0))
    for m in PCT_PATTERN.finditer(sentence):
        raw = m.group(0)
        tokens.add(raw)
        tokens.add(raw.rstrip("%"))
    for m in NUMBER_PATTERN.finditer(sentence):
        raw = m.group(0).replace(",", "")
        tokens.add(raw)
        try:
            f = float(raw)
            if f == int(f):
                tokens.add(str(int(f)))
            else:
                tokens.add(str(f))
                tokens.add(f"{f:.1f}".rstrip("0").rstrip("."))
        except ValueError:
            pass
    return {t for t in tokens if t}


def _fmt_trend(tr: dict[str, Any]) -> str:
    parts = [
        f"direction={tr.get('direction')}",
        f"first={tr.get('first')}",
        f"last={tr.get('last')}",
        f"pct_change={tr.get('pct_change')}",
    ]
    if tr.get("peak") is not None:
        parts.append(f"peak={tr.get('peak')} @ {tr.get('peak_period')}")
    if tr.get("trough") is not None:
        parts.append(f"trough={tr.get('trough')}")
    return ", ".join(str(p) for p in parts)


def build_metric_reference_lines(
    metrics: dict[str, Any] | None,
    narrative_context: dict[str, Any] | None,
    matched_columns: dict[str, Any] | None,
    *,
    include_key_findings: bool = True,
) -> tuple[list[str], set[str]]:
    """
    Flatten bundle-style fields into searchable lines and entity names
    (column labels, categories, group names).

    When include_key_findings is False (audit matching), key_findings lines are
    omitted so claims are not matched against narrative prose duplicated in the bundle.
    """
    lines: list[str] = []
    entities: set[str] = set()
    m = metrics or {}

    dr = m.get("date_range") or {}
    if isinstance(dr, dict) and (dr.get("min") is not None or dr.get("max") is not None):
        lines.append(f"date_range: min={dr.get('min')} max={dr.get('max')}")

    if m.get("total_categories") is not None:
        lines.append(f"total_categories: {m.get('total_categories')}")
    if m.get("total_periods") is not None:
        lines.append(f"total_periods: {m.get('total_periods')}")
    if m.get("total_points") is not None:
        lines.append(f"total_points: {m.get('total_points')}")

    for col, tr in (m.get("trend") or {}).items():
        if not isinstance(tr, dict):
            continue
        lines.append(f"trend[{col}]: {_fmt_trend(tr)}")
        for gname, gt in (tr.get("groups") or {}).items():
            if isinstance(gt, dict):
                lines.append(
                    f"trend[{col}].groups[{gname}]: direction={gt.get('direction')}, "
                    f"first={gt.get('first')}, last={gt.get('last')}, "
                    f"pct_change={gt.get('pct_change')}"
                )
            if isinstance(gname, str):
                entities.add(gname.lower())
        comp = tr.get("latest_period_comparison")
        if isinstance(comp, dict):
            vals = comp.get("values") or {}
            if isinstance(vals, dict):
                for grp, val in vals.items():
                    lines.append(f"latest_period_comparison[{col}] {grp}={val}")
                    if isinstance(grp, str):
                        entities.add(grp.lower())

    for col, stats in (m.get("summary_stats") or {}).items():
        if not isinstance(stats, dict):
            continue
        lines.append(
            f"summary_stats[{col}]: mean={stats.get('mean')}, min={stats.get('min')}, "
            f"max={stats.get('max')}, total={stats.get('total')}"
        )

    for col, ranking in (m.get("rankings") or {}).items():
        if not isinstance(ranking, list):
            continue
        for item in ranking:
            if not isinstance(item, dict):
                continue
            cat = item.get("category")
            val = item.get("value")
            lines.append(f"rankings[{col}]: {cat} = {val}")
            if isinstance(cat, str):
                entities.add(cat.lower())

    bb = m.get("bounding_box") or {}
    if isinstance(bb, dict) and bb:
        lines.append(
            f"bounding_box: lat [{bb.get('min_lat')}, {bb.get('max_lat')}], "
            f"lon [{bb.get('min_lon')}, {bb.get('max_lon')}]"
        )

    for d in m.get("category_distribution") or []:
        if isinstance(d, dict):
            lines.append(f"category_distribution: {d.get('category')} count={d.get('count')}")
            if isinstance(d.get("category"), str):
                entities.add(str(d["category"]).lower())

    nc = narrative_context or {}
    if include_key_findings:
        for i, finding in enumerate(nc.get("key_findings") or []):
            lines.append(f"key_findings[{i}]: {finding}")

    mc = matched_columns or {}
    if isinstance(mc, dict):
        for role, colname in mc.items():
            lines.append(f"matched_columns: {role} -> {colname}")
            if isinstance(colname, str):
                entities.add(colname.lower())
            if isinstance(role, str):
                entities.add(role.lower())

    dom = m.get("domain_extensions")
    if isinstance(dom, dict) and dom:
        try:
            dumped = json.dumps(dom, ensure_ascii=False)
            lines.append(f"domain_extensions: {dumped[:12000]}{'…' if len(dumped) > 12000 else ''}")
        except (TypeError, ValueError):
            lines.append(f"domain_extensions: {dom!r}")

    return lines, entities


def narrative_was_llm_generated(resp: dict[str, Any]) -> bool:
    """False for build_without_narrative / preview packages (provenance.llm_provider == none)."""
    prov = resp.get("provenance")
    if not isinstance(prov, dict):
        return False
    lp = str(prov.get("llm_provider") or "").strip().lower()
    if lp in ("", "none"):
        return False
    return True


def resolve_bundle_parts(record: dict[str, Any]) -> tuple[dict, dict | None, dict, str]:
    """metrics dict, narrative_context or None, matched_columns, template_type."""
    resp = record.get("response")
    if not isinstance(resp, dict):
        return {}, None, {}, ""
    bundle = resp.get("evidence_bundle")
    dataset = resp.get("dataset") or {}
    if isinstance(bundle, dict):
        metrics = bundle.get("metrics") or {}
        if not isinstance(metrics, dict):
            metrics = {}
        nc = bundle.get("narrative_context")
        if nc is not None and not isinstance(nc, dict):
            nc = {}
        mc = bundle.get("matched_columns")
        if not isinstance(mc, dict):
            mc = {}
        tmpl = bundle.get("template_type") or dataset.get("template_type") or ""
        return metrics, nc, mc, str(tmpl)
    metrics = {}
    nc = None
    mc = dataset.get("matched_columns") or {}
    if not isinstance(mc, dict):
        mc = {}
    tmpl = str(dataset.get("template_type") or "")
    return metrics, nc, mc, tmpl


_STOPWORDS = frozenset(
    {
        "the",
        "and",
        "for",
        "from",
        "with",
        "that",
        "this",
        "have",
        "been",
        "were",
        "was",
        "are",
        "has",
        "but",
        "not",
        "its",
        "their",
        "than",
        "into",
        "only",
        "also",
        "over",
        "such",
        "each",
        "most",
        "some",
        "between",
        "across",
        "chart",
        "data",
        "automated",
        "analysis",
    }
)


def _match_sentence_to_metric_lines(sentence: str, metric_lines: list[str]) -> tuple[list[str], bool]:
    toks = extract_number_tokens(sentence)
    matches: list[str] = []
    seen: set[str] = set()
    compact_blob = "\n".join(metric_lines)
    line_compacts = [ln.replace(",", "") for ln in metric_lines]

    if toks:
        for line, line_compact in zip(metric_lines, line_compacts):
            for t in toks:
                if not t:
                    continue
                t_norm = t.replace(",", "")
                if t_norm in line or t_norm in line_compact:
                    if line not in seen:
                        seen.add(line)
                        matches.append(line)
                    break
        if not matches and toks:
            for t in toks:
                if len(t) == 4 and t.isdigit() and t in compact_blob.replace(",", ""):
                    matches.append(
                        f"(year {t} appears in evidence_bundle metric lines — verify)"
                    )
                    break

    if not matches:
        words = re.findall(r"[\wæþðáéíóúýþöäüß·]+", sentence, re.I)
        keywords = {w.lower() for w in words if len(w) >= 4 and w.lower() not in _STOPWORDS}
        for line in metric_lines:
            ll = line.lower()
            hits = [kw for kw in keywords if kw in ll]
            if hits:
                if line not in seen:
                    seen.add(line)
                    matches.append(line)
                if len(matches) >= 10:
                    break

    return matches, bool(matches)


def evidence_reference_for_sentence(
    sentence: str,
    bundle_metric_lines: list[str],
) -> tuple[str, bool]:
    """Match claim tokens only to flattened evidence_bundle metric lines."""
    if not bundle_metric_lines:
        return (
            "NO MATCH FOUND — no evidence_bundle metric lines in JSON (missing or empty bundle)",
            False,
        )
    matches, ok = _match_sentence_to_metric_lines(sentence, bundle_metric_lines)
    if ok:
        return (
            " | ".join(matches[:12]) + (" | …" if len(matches) > 12 else ""),
            True,
        )
    return ("NO MATCH FOUND - check manually", False)


def run_id_from_filename(path: Path) -> str:
    stem = path.stem
    if stem.endswith("_query"):
        return stem[: -len("_query")]
    return stem


def load_sorted_corpus_json_files() -> list[Path]:
    paths = sorted(CORPUS_DIR.glob("run_*_query.json"))

    def sort_key(p: Path) -> tuple:
        rid = run_id_from_filename(p)
        m = re.match(r"^run_(\d+)([a-z]?)$", rid, re.I)
        if m:
            return (int(m.group(1)), m.group(2).lower())
        return (10**9, rid)

    return sorted(paths, key=sort_key)


def narrative_claim_texts(resp: dict[str, Any]) -> list[str]:
    texts: list[str] = []
    h = (resp.get("headline") or "").strip()
    if h:
        texts.extend(split_sentences(h) or [h])
    lede = (resp.get("lede") or "").strip()
    if lede:
        texts.extend(split_sentences(lede))
    for block in resp.get("story_blocks") or []:
        if not isinstance(block, dict):
            continue
        btype = (block.get("type") or "narrative").lower()
        if btype == "callout":
            hv = (block.get("highlight_value") or "").strip()
            hl = (block.get("highlight_label") or "").strip()
            if hv or hl:
                texts.append(" — ".join(x for x in (hv, hl) if x))
            body = (block.get("body") or "").strip()
            if body:
                texts.extend(split_sentences(body))
        else:
            body = (block.get("body") or "").strip()
            if body:
                texts.extend(split_sentences(body))
    dn = (resp.get("data_note") or "").strip()
    if dn:
        texts.extend(split_sentences(dn))
    return [t for t in texts if t.strip()]


def augment_entities_from_response(resp: dict[str, Any], entities_lower: set[str]) -> None:
    """Add dataset title tokens so place / entity names can be tagged as Entity."""
    name = (resp.get("dataset_name") or "").strip()
    if not name:
        return
    for word in re.findall(r"[\wæþðáéíóúýþöäüß]+", name, re.I):
        w = word.lower()
        if len(w) >= 4:
            entities_lower.add(w)


def clear_sheet_data(ws: Worksheet, first_row: int, last_col: int) -> None:
    for r in range(first_row, ws.max_row + 1):
        for c in range(1, last_col + 1):
            ws.cell(r, c).value = None


def extend_claim_audit_validations(ws: Worksheet, max_row: int = 8000) -> None:
    for dv in ws.data_validations.dataValidation:
        if not dv.sqref:
            continue
        s = str(dv.sqref)
        parts = []
        for rng in s.split():
            if ":" in rng:
                col_start, row_start = re.match(r"^([A-Z]+)(\d+)", rng.split(":")[0]).groups()
                col_end = rng.split(":")[1]
                col_end_letters = re.match(r"^([A-Z]+)", col_end).group(1)
                parts.append(f"{col_start}2:{col_end_letters}{max_row}")
            else:
                parts.append(rng)
        dv.sqref = " ".join(parts)


def main() -> None:
    if not TEMPLATE_XLSX.is_file():
        raise SystemExit(f"Template not found: {TEMPLATE_XLSX}")

    wb = load_workbook(TEMPLATE_XLSX)
    ws_claim = wb[CLAIM_SHEET]
    ws_sum = wb[SUMMARY_SHEET]
    ws_pipe = wb[PIPELINE_SHEET]

    clear_sheet_data(ws_claim, 2, 9)
    clear_sheet_data(ws_sum, 2, 15)
    clear_sheet_data(ws_pipe, 2, 14)

    files = load_sorted_corpus_json_files()
    if not files:
        raise SystemExit(f"No run_*_query.json files under {CORPUS_DIR}")

    claim_row = 2
    run_number = 0

    for path in files:
        run_number += 1
        record = json.loads(path.read_text(encoding="utf-8"))
        run_id = run_id_from_filename(path)
        q = str(record.get("query") or "")
        resp = record.get("response")
        has_bundle = isinstance(resp, dict) and isinstance(resp.get("evidence_bundle"), dict)

        metrics, nc, mc, tmpl = resolve_bundle_parts(record)
        if has_bundle:
            metric_lines, entity_set = build_metric_reference_lines(
                metrics, nc, mc, include_key_findings=False
            )
        else:
            metric_lines, entity_set = [], set()

        if isinstance(resp, dict):
            ds = resp.get("dataset") or {}
            if isinstance(ds, dict):
                for colname in (ds.get("matched_columns") or {}).values():
                    if isinstance(colname, str) and len(colname.strip()) >= 2:
                        entity_set.add(colname.strip().lower())

        entities_lower = {e.lower() for e in entity_set if e}

        if not isinstance(resp, dict):
            texts = ["[No successful response — see JSON error fields]"]
        else:
            augment_entities_from_response(resp, entities_lower)
            texts = narrative_claim_texts(resp)

        claim_count = len(texts)

        for i, text in enumerate(texts, start=1):
            ev, ev_ok = evidence_reference_for_sentence(text, metric_lines)
            ctype = classify_claim_type(text, entities_lower, evidence_matched=ev_ok)
            verdict = VERDICT_AUTO_MATCH if ev_ok else None
            if ev_ok:
                note_claim = NOTE_AUTO_MATCH
            elif not isinstance(resp, dict):
                note_claim = "No API payload."
            elif not has_bundle:
                note_claim = NOTE_NO_BUNDLE
            else:
                note_claim = None

            is_first = i == 1
            ws_claim.cell(claim_row, 1, run_id if is_first else None)
            ws_claim.cell(claim_row, 2, q if is_first else None)
            ws_claim.cell(claim_row, 3, tmpl if is_first else None)
            ws_claim.cell(claim_row, 4, i)
            ws_claim.cell(claim_row, 5, text)
            ws_claim.cell(claim_row, 6, ctype)
            ws_claim.cell(claim_row, 7, ev)
            ws_claim.cell(claim_row, 8, verdict)
            ws_claim.cell(claim_row, 9, note_claim)
            claim_row += 1

        # Narrative Summary
        sum_r = run_number + 1
        ws_sum.cell(sum_r, 1, run_id)
        ws_sum.cell(sum_r, 2, q)
        ws_sum.cell(sum_r, 3, tmpl)
        ws_sum.cell(sum_r, 4, run_number)
        ws_sum.cell(sum_r, 7, claim_count)
        ws_sum.cell(sum_r, 11, f"=IF(G{sum_r}=0,\"\",H{sum_r}/G{sum_r})")
        llm_ok = isinstance(resp, dict) and narrative_was_llm_generated(resp)
        gen_time = record.get("response_time_seconds")
        if llm_ok:
            ws_sum.cell(sum_r, 14, gen_time)
        else:
            ws_sum.cell(sum_r, 14, None)
        if isinstance(resp, dict) and not llm_ok:
            ws_sum.cell(
                sum_r,
                15,
                "HTTP 200 with preview/fallback package (no LLM narrative). "
                "Generation time left blank so averages reflect successful LLM runs only.",
            )
        elif not isinstance(resp, dict):
            ws_sum.cell(sum_r, 15, "No response payload — see JSON error fields.")
        else:
            ws_sum.cell(sum_r, 15, None)

        # Pipeline Success (A–N): see template — K = LLM narrative, L = full pipeline
        ds_label = ""
        if isinstance(resp, dict):
            name = (resp.get("dataset_name") or "").strip()
            ds = resp.get("dataset") or {}
            did = ""
            if isinstance(ds, dict):
                did = str(ds.get("dataset_id") or "")
            if name and did:
                ds_label = f"{name} ({did})"
            elif did:
                ds_label = did
            elif name:
                ds_label = name
        pipe_r = run_number + 1
        ws_pipe.cell(pipe_r, 1, run_id)
        ws_pipe.cell(pipe_r, 2, q)
        ws_pipe.cell(pipe_r, 3, ds_label or "")
        ok_payload = isinstance(resp, dict)
        for col in (4, 5, 6):
            ws_pipe.cell(pipe_r, col, "Yes" if ok_payload else "No")
        ws_pipe.cell(pipe_r, 7, tmpl if ok_payload else "")
        has_charts = (
            ok_payload
            and isinstance(resp.get("visualizations"), list)
            and len(resp["visualizations"]) > 0
        )
        ws_pipe.cell(pipe_r, 8, "Yes" if ok_payload else "No")
        ws_pipe.cell(pipe_r, 9, "Yes" if has_charts else "No")
        ws_pipe.cell(pipe_r, 10, "Yes" if has_charts else "No")
        ws_pipe.cell(pipe_r, 11, "Yes" if llm_ok else "No")
        ws_pipe.cell(pipe_r, 12, "Yes" if llm_ok else "No")
        if not ok_payload:
            ws_pipe.cell(pipe_r, 13, "No API payload / request error")
        elif not llm_ok:
            ws_pipe.cell(
                pipe_r,
                13,
                "LLM narrative not produced (preview / template analysis only)",
            )
        else:
            ws_pipe.cell(pipe_r, 13, None)
        ws_pipe.cell(pipe_r, 14, None)

    extend_claim_audit_validations(ws_claim, max_row=max(8000, claim_row + 100))

    wb.save(OUTPUT_XLSX)
    print(f"Wrote {OUTPUT_XLSX} ({len(files)} runs, {claim_row - 2} claim rows).")


if __name__ == "__main__":
    main()
